from django.shortcuts import render

# Create your views here.
from testapp.models import Report
from django.http import HttpResponse
from django.views.generic import View
from django.db.models.signals import post_save
from django.dispatch import receiver

from django.http import HttpResponse
from django.views.generic import View
from testapp.utils import render_to_pdf
import openpyxl
import json
import pdfkit



from weasyprint import HTML
from django.template import Template, Context
from django.template.loader import render_to_string, get_template
import json





def input(request):
    return render(request,'testapp/forms.html')

#
#
# def index(request):
#
#     if "GET" == request.method:
#         return render(request, 'testapp/index.html', {})
#
#     else:
#         excel_file = request.FILES["excel_file"]
#
#         wb = openpyxl.load_workbook(excel_file,data_only = True)
#
#         # getting a particular sheet by name out of many sheets
#         worksheet = wb["Sheet1"]
#
#
#         # worksheet = wb["Sheet1"]
#         print(worksheet)
#
#
# #################################  reading excel data        #######################################
#
#         excel_data = list()
#         sd = {}
#         # iterating over the rows and column
#         # getting value from each cell in row
#         for row in worksheet.iter_rows(min_col=2,max_col =4,min_row =4,max_row =8):
#             row_data = list()
#             sd['name'] = row[0].value
#             sd['phone'] = row[1].value
#             sd['address'] = row[2].value
#             r = Report(**sd)
#             r.save()
#
#         report = Report.objects.all()
#         pdf = render_to_pdf('pdf/details.html', {'report': report})
#         return HttpResponse(pdf, content_type='application/pdf')
#
#
#
#
# def index(request):
#
#     if "GET" == request.method:
#         return render(request, 'testapp/index.html', {})
#
#     else:
#         excel_file = request.FILES["excel_file"]
#
#         wb = openpyxl.load_workbook(excel_file,data_only = True)
#
#         # getting a particular sheet by name out of many sheets
#         worksheet = wb["Sheet1"]
#
#
#         # worksheet = wb["Sheet1"]
#         print(worksheet)
#
#
# #################################  reading excel data        #######################################
#
#         excel_data = list()
#         sd = {}
#         # iterating over the rows and column
#         # getting value from each cell in row
#         for row in worksheet.iter_rows(min_col=2,max_col =4,min_row =4,max_row =8):
#             row_data = list()
#             sd['name'] = row[0].value
#             sd['phone'] = row[1].value
#             sd['address'] = row[2].value
#             r = Report(**sd)
#             r.save()
#
#         report = Report.objects.all()
#         pdf = render_to_pdf('pdf/details.html', {'report': report})
#         return HttpResponse(pdf, content_type='application/pdf')
#
#


##################################################################################################################


#
# def index(request):
#
#     if "GET" == request.method:
#         return render(request, 'testapp/index.html', {})
#
#     else:
#         excel_file = request.FILES["excel_file"]
#
#         wb = openpyxl.load_workbook(excel_file,data_only = True)
#
#         # getting a particular sheet by name out of many sheets
#         worksheet = wb["Sheet1"]
#
#
#         # worksheet = wb["Sheet1"]
#         print(worksheet)
#
#
# #################################  reading excel data        #######################################
#
#         excel_data = list()
#         my_dict = {}
#
#
#         for col in worksheet.iter_cols(min_col=2,max_col=4,min_row = 4 ,max_row =8):
#             for cell in col:
#                 if cell.value is None:
#                     continue
#
#
#
#
#         report = Report.objects.all()
#         pdf = render_to_pdf('pdf/details.html', {'report': report})
#         return HttpResponse(pdf, content_type='application/pdf')


def index(request):

    if "GET" == request.method:
        return render(request, 'testapp/index.html', {})

    else:
        excel_file = request.FILES["excel_file"]

        wb = openpyxl.load_workbook(excel_file,data_only =True)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]


        # worksheet = wb["Sheet1"]
        print(worksheet)


#################################  reading excel data        #######################################

        # excel_data = list()
        my_dict = {}

        # iterating over the rows and column
        # getting value from each cell in



        my_dict[str(worksheet['K4'].value)] = str(worksheet['L4'].value)
        my_dict[str(worksheet['K6'].value)] = int(worksheet['L6'].value)
        my_dict[str(worksheet['K7'].value)] = int(worksheet['L7'].value)
        my_dict[str(worksheet['K8'].value)] = int(worksheet['L8'].value)

        print(my_dict)
        my_dict1 ={}
        my_dict1[str(worksheet['K8'].value)] = int(worksheet['L8'].value)

        # # data = json.dumps(my_dict)
        # #
        # # print(data)
        # # p = pdfkit.from_string(data,'pratik.pdf')
        #
        # #return HttpResponse('pratik.pdf',content_type='application/pdf')
        # return render('pdf.html', p)
        # # # return HttpResponse(p)
        # #
        # # return HttpResponse(p, content_type='application/pdf')



        pdf = render_to_pdf('pdf/details.html',{'data': my_dict.items(),'data1':my_dict1.items()})
        return HttpResponse(pdf, content_type='application/pdf')














































            #     row_data.append(str(cell.value))
            # excel_data.append(row_data)

        # return render(request,'testapp/forms.html')


# def show(request):
#     report = Report.objects.all()
#     return render(request,'testapp/details1.html',{'report':report})
#
#
# def show(request):
#     # qs = Report.objects.all()
#     return render(request,'pratikapp/forms.html')

#
# @receiver(post_save,sender = Report)
# class GeneratePdf(View):
#     def get(self, request, *args, **kwargs):
#         if kwargs.get('created',False):
#             report = Report.objects.last()
#
#         pdf = render_to_pdf('pdf/details.html', {'report':report})
#         return HttpResponse(pdf, content_type='application/pdf')



#
# @receiver(post_save, sender = User)
# def ensure_profile_exists(sender, **kwargs):
#     if kwargs.get('created', False):


            #     row_data.append(str(cell.value))
            # excel_data.append(row_data)

        # return render(request,'testapp/forms.html')


# def show(request):
#     report = Report.objects.all()
#     return render(request,'testapp/details1.html',{'report':report})
#
#
# def show(request):
#     # qs = Report.objects.all()
#     return render(request,'pratikapp/forms.html')

#
# @receiver(post_save,sender = Report)
# class GeneratePdf(View):
#     def get(self, request, *args, **kwargs):
#         if kwargs.get('created',False):
#             report = Report.objects.last()
#
#         pdf = render_to_pdf('pdf/details.html', {'report':report})
#         return HttpResponse(pdf, content_type='application/pdf')



#
# @receiver(post_save, sender = User)
# def ensure_profile_exists(sender, **kwargs):
#     if kwargs.get('created', False):
#         Report.objects.get_or_create(user=kwargs.get('instance'))